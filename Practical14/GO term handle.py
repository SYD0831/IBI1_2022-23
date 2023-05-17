import xml.dom.minidom as minidom
from openpyxl import Workbook
# Parse the XML file
dom = minidom.parse('go_obo.xml')
root = dom.documentElement


# Initialize Excel workbook and sheet
workbook = Workbook()
sheet = workbook.active
sheet.title = 'GO Terms'

# Set column headers
headers = ['GO ID', 'Term Name', 'Definition', 'Child Nodes']
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Find 'autophagosome' related gene ontology terms and their child nodes
row_num = 2  # Start from row 2
terms = root.getElementsByTagName('term')

#We need to find all child nodes of one node and its child nodes until it reaches the bottom.
#So we can get the number through calculating the number of parent nodes of the child nodes.
#This function can get the id  of parent nodes
nodefamily = {}
result_all = {}
def findnodes(node_id, parent_ids):
    mediate_ids = nodefamily[node_id]
    for mediate_id in mediate_ids:
        parent_ids.add(mediate_id) 
        findnodes(mediate_id, parent_ids)

#We need to get all nodes of each term
#The information will be stored in the nodefamily.
for term in terms:
    node_id = term.getElementsByTagName("id")[0].childNodes[0].data
    mediate_ids = []
    for is_a in term.getElementsByTagName("is_a"):
        mediate_ids.append(is_a.childNodes[0].data)
    nodefamily[node_id] = mediate_ids
    result_all[node_id] = 0

#Now we need to calculate all parent nodes of each child node.
#This equals the number of all child nodes of each parent node.
for key in nodefamily.keys():
    parent_ids = set()
    findnodes(key, parent_ids)
    for parent_id in parent_ids:
        result_all[parent_id] += 1

        
for term in terms:
    go_id = term.getElementsByTagName('id')[0].childNodes[0].data
    term_name = term.getElementsByTagName('name')[0].childNodes[0].data
    defstr = term.getElementsByTagName('defstr')[0].childNodes[0].data

    # Check if 'autophagosome' is present in the definition string
    if 'autophagosome' in defstr.lower():
        child_count = result_all[go_id]
        # Write the information to the spreadsheet
        sheet.cell(row=row_num, column=1, value=go_id)
        sheet.cell(row=row_num, column=2, value=term_name)
        sheet.cell(row=row_num, column=3, value=defstr)
        sheet.cell(row=row_num, column=4, value=child_count)
        row_num += 1

# Save the workbook
workbook.save('go_terms.xlsx')

